from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
import sys

from openpyxl import load_workbook
from openpyxl.styles import Alignment


COVER_SHEETS = {"Guideline", "Cover", "MethodList", "Statistics"}


def copy_row_style(ws, source_row: int, target_row: int, start_col: int = 1, end_col: int = 6) -> None:
    for col in range(start_col, end_col + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def find_executed_dates(ws) -> list[datetime]:
    for row in range(1, min(ws.max_row, 120) + 1):
        label = ws.cell(row=row, column=2).value
        if isinstance(label, str) and label.strip() == "Executed Date":
            dates: list[datetime] = []
            for col in range(6, ws.max_column + 1):
                value = ws.cell(row=row, column=col).value
                if isinstance(value, datetime):
                    dates.append(value)
            return dates
    return []


def sheet_method_name(ws) -> str:
    for cell in ("L1", "Q1", "R1", "M1"):
        value = ws[cell].value
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ws.title


def build_change_rows(wb) -> list[dict[str, object]]:
    grouped: dict[str, dict[str, object]] = {}

    for ws in wb.worksheets:
        if ws.title in COVER_SHEETS:
            continue

        creator = str(ws["C2"].value or "").strip()
        module = str(ws["C1"].value or "").strip()
        method = sheet_method_name(ws)
        executed_dates = find_executed_dates(ws)

        if creator not in grouped:
            grouped[creator] = {
                "creator": creator,
                "modules": [],
                "sheets": [],
                "methods": [],
                "executed_dates": [],
            }

        entry = grouped[creator]
        if module and module not in entry["modules"]:
            entry["modules"].append(module)
        entry["sheets"].append(ws.title)
        entry["methods"].append(method)
        entry["executed_dates"].extend(executed_dates)

    rows: list[dict[str, object]] = []
    for creator, entry in grouped.items():
        executed_dates = sorted({dt.date() for dt in entry["executed_dates"]})
        effective_date = executed_dates[-1] if executed_dates else None

        modules = ", ".join(entry["modules"])
        sheets = entry["sheets"]
        if len(sheets) == 1:
            reference = sheets[0]
        elif len(sheets) == 2:
            reference = f"{sheets[0]}; {sheets[1]}"
        else:
            reference = f"{sheets[0]} to {sheets[-1]}"

        description = (
            f"Recorded {len(sheets)} UT function sheets created by {creator} "
            f"for modules {modules} based on tester and executed-date information "
            f"already entered in each UT tab."
        )

        rows.append(
            {
                "effective_date": effective_date,
                "version": "1.0",
                "change_item": f"{creator} UT records",
                "change_type": "A",
                "description": description,
                "reference": reference,
            }
        )

    rows.sort(
        key=lambda item: (
            item["effective_date"] is None,
            item["effective_date"] or datetime.min.date(),
            str(item["change_item"]).lower(),
        )
    )
    return rows


def apply_change_rows(cover_ws, rows: list[dict[str, object]]) -> None:
    start_row = 11
    for offset, row_data in enumerate(rows):
        row = start_row + offset
        copy_row_style(cover_ws, 11, row)

        cover_ws[f"A{row}"] = row_data["effective_date"]
        cover_ws[f"B{row}"] = row_data["version"]
        cover_ws[f"C{row}"] = row_data["change_item"]
        cover_ws[f"D{row}"] = row_data["change_type"]
        cover_ws[f"E{row}"] = row_data["description"]
        cover_ws[f"F{row}"] = row_data["reference"]

        cover_ws[f"A{row}"].number_format = "d-mmm-yy"
        cover_ws[f"B{row}"].number_format = "@"
        cover_ws[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        cover_ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="top")
        cover_ws[f"E{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        cover_ws[f"F{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        cover_ws.row_dimensions[row].height = 42


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: update_record_of_changes.py <input.xlsx> <output.xlsx>", file=sys.stderr)
        return 1

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    wb = load_workbook(input_path)
    cover_ws = wb["Cover"]
    rows = build_change_rows(wb)
    apply_change_rows(cover_ws, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
